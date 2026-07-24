"""Build one run-local KG from the immutable base graph and prepared inputs."""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import tempfile
import time
from collections import Counter, defaultdict
from collections.abc import Callable, Iterable, Mapping, Sequence
from contextlib import ExitStack
from dataclasses import dataclass
from importlib import metadata
from pathlib import Path
from typing import Any, TextIO

from lipid_screening_agent import __version__
from lipid_screening_agent.artifacts import (
    ArtifactManifest,
    NodeResult,
    NodeStatus,
    load_artifact_manifest,
    verify_artifact_manifest,
)
from lipid_screening_agent.cli import (
    CommonRunnerArguments,
    add_common_runner_arguments,
    load_common_runner_environment,
)
from lipid_screening_agent.config.models import DiseaseConfig, KGConstructionConfig
from lipid_screening_agent.runtime import (
    EnvironmentError,
    InputError,
    OutputContractError,
    PathSafetyError,
    ResourceError,
    RunContext,
    atomic_write_json,
    ensure_within,
    resolve_run_relative,
    sha256_file,
)
from lipid_screening_agent.runtime.execution import NodeExecution, execute_node

from ..input_prepare._common import add_execution_identity_arguments, execution_identity

NODE_ID = "kg_construct_graph"
OUTPUT_ROOT = "artifacts/kg/construction"
SCHEMA_VERSION = "1.0"
DRUG_TARGET_PROVIDERS = frozenset(
    {
        ("netinfer_merge_targets", "artifacts/netinfer/drug_targets.json"),
        ("import_drug_targets", "artifacts/targets/drug_targets.json"),
    }
)

NODE_COLUMNS = (
    "node_index",
    "node_id",
    "node_type",
    "node_name",
    "node_source",
    "source_ids",
    "description",
    "node_type_raw",
)
EDGE_COLUMNS = (
    "edge_id",
    "x_index",
    "y_index",
    "x_id",
    "y_id",
    "x_type",
    "y_type",
    "relation",
    "relation_raw",
    "edge_type",
    "source_db",
    "evidence",
    "score",
    "rank",
    "original_relation",
)
KG_COLUMNS = (
    "x_type",
    "x_id",
    "x_index",
    "x_name",
    "x_source",
    "relation",
    "y_type",
    "y_id",
    "y_index",
    "y_name",
    "y_source",
    "source_db",
    "evidence",
    "score",
    "rank",
    "original_relation",
    "edge_type",
    "edge_id",
)
MATCH_REPORT_COLUMNS = (
    "library_id",
    "input_smiles",
    "canonical_smiles",
    "output_node_id",
    "is_new_node",
    "match_type",
    "matched_base_node_id",
    "matched_base_node_name",
)
INVALID_SMILES_COLUMNS = ("ID", "SMILES", "reason")
UNMAPPED_TARGET_COLUMNS = (
    "gene_symbol",
    "occurrence_count",
    "compound_ids",
    "reason",
)

OUTPUT_FILES = {
    "kg_nodes": "node.csv",
    "kg_edges": "edges.csv",
    "kg_graph": "kg.csv",
    "kg_construction_manifest": "manifest.json",
    "kg_drug_match_report": "drug_smiles_match_report.tsv",
    "kg_invalid_smiles": "invalid_smiles.tsv",
    "kg_unmapped_targets": "unmapped_target_symbols.tsv",
}


@dataclass(slots=True)
class CompoundRecord:
    library_id: str
    smiles: str
    canonical_smiles: str
    name: str
    cas: str
    formula: str
    molwt: str
    node_id: str = ""
    is_new_node: bool = True
    match_type: str = "new_user_node"
    matched_base_node_id: str = ""


@dataclass(frozen=True, slots=True)
class TargetRecord:
    symbol: str
    uniprot_id: str
    evidence: str
    score: str
    rank: str


@dataclass(frozen=True, slots=True)
class EdgeCandidate:
    x_id: str
    y_id: str
    relation: str
    relation_raw: str
    edge_type: str
    source_db: str
    evidence: str = ""
    score: str = ""
    rank: str = ""
    original_relation: str = ""


def _clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.casefold() == "nan" else text


def _symbol(value: object) -> str:
    return _clean(value).upper()


def _gene_id(value: object) -> str:
    text = _clean(value).strip(";")
    return text.split(";", 1)[0].strip()


def _protein_id(entrez_id: object) -> str:
    cleaned = _gene_id(entrez_id)
    return f"protein:ncbi:{cleaned}" if cleaned else ""


def _append_source_ids(*parts: object) -> str:
    seen: set[str] = set()
    values: list[str] = []
    for part in parts:
        for value in _clean(part).split(";"):
            value = value.strip()
            if value and value.casefold() != "nan" and value not in seen:
                seen.add(value)
                values.append(value)
    return ";".join(values)


def _require_columns(fieldnames: Sequence[str] | None, required: Sequence[str], label: str) -> None:
    observed = set(fieldnames or ())
    missing = [column for column in required if column not in observed]
    if missing:
        raise InputError(
            f"{label} is missing required columns",
            details={"missing_columns": missing, "observed_columns": list(fieldnames or ())},
        )


def _package_version(distribution: str) -> str:
    try:
        return metadata.version(distribution)
    except metadata.PackageNotFoundError:
        return "not_installed"
    except Exception:
        return "unknown"


def _resolve_output(context: RunContext, filename: str) -> Path:
    expected_root = context.resolve_run_relative(OUTPUT_ROOT)
    if context.output_dir != expected_root:
        raise OutputContractError(
            "KG output_dir must be the contracted artifacts/kg/construction directory",
            details={
                "configured_output_dir": str(context.output_dir),
                "expected_output_dir": str(expected_root),
            },
        )
    try:
        return resolve_run_relative(expected_root, filename)
    except (OSError, PathSafetyError) as exc:
        raise OutputContractError("unsafe KG output path", details={"filename": filename}) from exc


def _resolve_run_file(context: RunContext, value: str | Path, *, label: str) -> Path:
    candidate = Path(value)
    try:
        if candidate.is_absolute():
            resolved = ensure_within(candidate, context.run_dir)
        else:
            resolved = context.resolve_run_relative(candidate.as_posix(), must_exist=True)
    except (OSError, PathSafetyError) as exc:
        raise InputError(
            f"{label} is missing or outside the run workspace",
            details={"path": str(value)},
        ) from exc
    if not resolved.exists() or not resolved.is_file() or resolved.is_symlink():
        raise InputError(
            f"{label} must be an existing regular non-symlink file",
            details={"path": str(resolved)},
        )
    return resolved


def _resolve_optional_run_file(
    context: RunContext, value: str | Path | None, *, label: str
) -> Path | None:
    return None if value is None else _resolve_run_file(context, value, label=label)


def _resolve_resource_file(context: RunContext, value: str | Path, *, label: str) -> Path:
    if context.resource_dir is None:
        raise ResourceError("a resource directory is required for KG construction")
    candidate = Path(value)
    try:
        if candidate.is_absolute():
            resolved = ensure_within(candidate, context.resource_dir)
        else:
            resolved = resolve_run_relative(
                context.resource_dir, candidate.as_posix(), must_exist=True
            )
    except (OSError, PathSafetyError) as exc:
        raise ResourceError(
            f"{label} is missing or outside the KG resource boundary",
            details={"path": str(value), "resource_dir": str(context.resource_dir)},
        ) from exc
    if not resolved.exists() or not resolved.is_file() or resolved.is_symlink():
        raise ResourceError(
            f"{label} must be an existing regular non-symlink file",
            details={"path": str(resolved)},
        )
    return resolved


def _resolve_drug_targets(
    context: RunContext,
    *,
    drug_targets_path: str | Path | None,
    drug_targets_manifest_path: str | Path | None,
) -> tuple[Path, ArtifactManifest | None]:
    if (drug_targets_path is None) == (drug_targets_manifest_path is None):
        raise InputError("supply exactly one of drug_targets_path or drug_targets_manifest_path")
    if drug_targets_path is not None:
        return _resolve_run_file(
            context, drug_targets_path, label="NetInfer drug_targets.json"
        ), None

    manifest_file = _resolve_run_file(
        context,
        drug_targets_manifest_path,  # type: ignore[arg-type]
        label="drug_targets artifact manifest",
    )
    try:
        manifest = load_artifact_manifest(manifest_file, run_root=context.run_dir)
        verify_artifact_manifest(manifest, run_root=context.run_dir)
    except Exception as exc:
        raise InputError(
            "drug_targets artifact manifest is invalid or stale",
            details={"manifest_path": str(manifest_file)},
        ) from exc
    provider = (manifest.producer_node_id, manifest.relative_path)
    if manifest.artifact_type != "drug_targets" or provider not in DRUG_TARGET_PROVIDERS:
        raise InputError(
            "artifact manifest is not from an approved drug_targets provider",
            details={
                "approved_providers": sorted(DRUG_TARGET_PROVIDERS),
                "observed": {
                    "artifact_type": manifest.artifact_type,
                    "producer_node_id": manifest.producer_node_id,
                    "relative_path": manifest.relative_path,
                },
            },
        )
    result_path = context.resolve_run_relative(
        f"artifacts/node_results/{manifest.producer_node_id}/{manifest.producer_task_id}.json",
        must_exist=True,
    )
    try:
        result_payload = json.loads(result_path.read_text(encoding="utf-8"))
        result = NodeResult.from_dict(result_payload)
    except Exception as exc:
        raise InputError("drug_targets producer NodeResult is missing or invalid") from exc
    if result.status not in {NodeStatus.SUCCEEDED, NodeStatus.CACHED}:
        raise InputError(
            "drug_targets producer did not finish successfully",
            details={"status": result.status.value},
        )
    if manifest.artifact_id not in result.outputs:
        raise InputError(
            "drug_targets manifest was not committed by the successful NodeResult",
            details={"artifact_id": manifest.artifact_id},
        )
    return context.resolve_run_relative(manifest.relative_path, must_exist=True), manifest


def _load_chemistry() -> tuple[Any, Any | None]:
    try:
        from rdkit import Chem
    except ImportError as exc:
        raise EnvironmentError(
            "RDKit is required for KG SMILES standardization", retryable=False
        ) from exc
    try:
        from rdkit.Chem.MolStandardize import rdMolStandardize
    except Exception:
        rdMolStandardize = None
    return Chem, rdMolStandardize


def _load_fingerprint(*, workers: int) -> Any:
    try:
        from skfp.fingerprints import PubChemFingerprint
    except (ImportError, ModuleNotFoundError) as exc:
        raise EnvironmentError(
            "scikit-fingerprints is required for PubChemFP substructure edges",
            details={"distribution": "scikit-fingerprints"},
            retryable=False,
        ) from exc
    try:
        return PubChemFingerprint(n_jobs=workers)
    except Exception as exc:
        raise EnvironmentError(
            "PubChemFingerprint could not be initialized",
            details={"workers": workers},
            retryable=False,
        ) from exc


def _standardize_smiles(smiles: object, *, chem: Any, standardizer: Any | None) -> str:
    text = _clean(smiles)
    if not text:
        return ""
    mol = chem.MolFromSmiles(text)
    if mol is None:
        return ""
    if standardizer is not None:
        try:
            mol = standardizer.Cleanup(mol)
        except Exception:
            pass
    try:
        return chem.MolToSmiles(mol, canonical=True, isomericSmiles=True)
    except Exception:
        return ""


def _read_nodes(path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    node_ids: set[str] = set()
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter="\t")
            _require_columns(reader.fieldnames, NODE_COLUMNS, "base node table")
            for row_number, raw in enumerate(reader, start=2):
                row = {column: _clean(raw.get(column)) for column in NODE_COLUMNS}
                node_id = row["node_id"]
                if not node_id:
                    raise ResourceError(
                        "base node table contains an empty node_id",
                        details={"row": row_number},
                    )
                if node_id in node_ids:
                    raise ResourceError(
                        "base node table contains duplicate node_id values",
                        details={"node_id": node_id, "row": row_number},
                    )
                node_ids.add(node_id)
                rows.append(row)
    except (InputError, ResourceError):
        raise
    except (OSError, UnicodeError, csv.Error) as exc:
        raise ResourceError(
            "base node table could not be read", details={"path": str(path)}
        ) from exc
    if not rows:
        raise ResourceError("base node table contains no rows")
    return rows


def _read_compounds(
    path: Path, *, chem: Any, standardizer: Any | None
) -> tuple[list[CompoundRecord], list[dict[str, str]]]:
    compounds: list[CompoundRecord] = []
    invalid: list[dict[str, str]] = []
    seen: set[str] = set()
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            _require_columns(reader.fieldnames, ("ID", "SMILES"), "normalized compounds")
            for row_number, row in enumerate(reader, start=2):
                library_id = _clean(row.get("ID"))
                smiles = _clean(row.get("SMILES"))
                if not library_id:
                    raise InputError(
                        "normalized compounds contain an empty ID",
                        details={"row": row_number},
                    )
                if library_id in seen:
                    raise InputError(
                        "normalized compounds contain duplicate IDs",
                        details={"ID": library_id, "row": row_number},
                    )
                seen.add(library_id)
                canonical = _standardize_smiles(smiles, chem=chem, standardizer=standardizer)
                if not canonical:
                    invalid.append({"ID": library_id, "SMILES": smiles, "reason": "invalid_smiles"})
                compounds.append(
                    CompoundRecord(
                        library_id=library_id,
                        smiles=smiles,
                        canonical_smiles=canonical,
                        name=_clean(row.get("name")) or _clean(row.get("Name")) or library_id,
                        cas=_clean(row.get("CAS")),
                        formula=_clean(row.get("Formula")),
                        molwt=_clean(row.get("MolWt")),
                    )
                )
    except InputError:
        raise
    except (OSError, UnicodeError, csv.Error) as exc:
        raise InputError(
            "normalized compounds could not be read", details={"path": str(path)}
        ) from exc
    if not compounds:
        raise InputError("normalized compounds contain no records")
    return compounds, invalid


def _read_disease_genes(path: Path) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    seen: set[str] = set()
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter="\t")
            _require_columns(reader.fieldnames, ("symbol", "entrez_id"), "normalized disease genes")
            for row in reader:
                entrez_id = _gene_id(row.get("entrez_id"))
                if not entrez_id or entrez_id in seen:
                    continue
                seen.add(entrez_id)
                rows.append((_symbol(row.get("symbol")), entrez_id))
    except InputError:
        raise
    except (OSError, UnicodeError, csv.Error) as exc:
        raise InputError(
            "normalized disease genes could not be read", details={"path": str(path)}
        ) from exc
    if not rows:
        raise InputError("normalized disease genes contain no usable Entrez IDs")
    return rows


def _mapping_from_rows(rows: Iterable[Sequence[object]], *, source_path: Path) -> dict[str, str]:
    iterator = iter(rows)
    header_row = next(iterator, None)
    if header_row is None:
        raise ResourceError("target mapping contains no rows")
    header = [_clean(value) for value in header_row]
    symbol_aliases = {"Gene symbol", "gene_symbol", "symbol"}
    entrez_aliases = {"Gene ID", "entrez_id", "gene_id"}
    symbol_indices = [index for index, value in enumerate(header) if value in symbol_aliases]
    entrez_indices = [index for index, value in enumerate(header) if value in entrez_aliases]
    if len(symbol_indices) != 1 or len(entrez_indices) != 1:
        raise ResourceError(
            "target mapping must contain one gene-symbol and one Entrez-ID column",
            details={"path": str(source_path), "columns": header},
        )
    symbol_index, entrez_index = symbol_indices[0], entrez_indices[0]
    mapping: dict[str, str] = {}
    ambiguous: dict[str, set[str]] = defaultdict(set)
    for row in iterator:
        if len(row) <= max(symbol_index, entrez_index):
            continue
        symbol = _symbol(row[symbol_index])
        entrez_id = _gene_id(row[entrez_index])
        if not symbol or not entrez_id:
            continue
        previous = mapping.get(symbol)
        if previous is not None and previous != entrez_id:
            ambiguous[symbol].update((previous, entrez_id))
        else:
            mapping[symbol] = entrez_id
    if ambiguous:
        raise ResourceError(
            "target mapping contains symbols assigned to multiple Entrez IDs",
            details={
                "ambiguous_symbols": {
                    key: sorted(values) for key, values in list(sorted(ambiguous.items()))[:20]
                }
            },
        )
    if not mapping:
        raise ResourceError("target mapping contains no usable symbol-to-Entrez records")
    return mapping


def _read_target_mapping(path: Path) -> dict[str, str]:
    if path.suffix.casefold() in {".xlsx", ".xlsm"}:
        try:
            import openpyxl
        except ImportError as exc:
            raise EnvironmentError(
                "openpyxl is required to read the NetInfer target workbook",
                retryable=False,
            ) from exc
        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
            if "Target information" not in workbook.sheetnames:
                raise ResourceError("target workbook is missing the 'Target information' sheet")
            sheet = workbook["Target information"]
            mapping = _mapping_from_rows(sheet.iter_rows(values_only=True), source_path=path)
            workbook.close()
            return mapping
        except (ResourceError, EnvironmentError):
            raise
        except Exception as exc:
            raise ResourceError(
                "target workbook could not be read", details={"path": str(path)}
            ) from exc
    delimiter = "," if path.suffix.casefold() == ".csv" else "\t"
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return _mapping_from_rows(csv.reader(handle, delimiter=delimiter), source_path=path)
    except ResourceError:
        raise
    except (OSError, UnicodeError, csv.Error) as exc:
        raise ResourceError(
            "target mapping table could not be read", details={"path": str(path)}
        ) from exc


def _read_drug_targets(path: Path, *, top_n: int) -> dict[str, list[TargetRecord]]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise InputError("NetInfer drug_targets.json could not be read") from exc
    if not isinstance(payload, dict):
        raise InputError("NetInfer drug_targets.json must be keyed by compound ID")
    output: dict[str, list[TargetRecord]] = {}
    for compound_id, info in payload.items():
        if not isinstance(compound_id, str) or not compound_id.strip():
            raise InputError("drug_targets.json contains an empty compound ID")
        if not isinstance(info, dict) or not isinstance(info.get("targets"), list):
            raise InputError(
                "each drug_targets.json value must contain a targets list",
                details={"ID": compound_id},
            )
        known: list[tuple[int, Mapping[str, Any]]] = []
        predicted: list[tuple[int, int, Mapping[str, Any]]] = []
        for position, target in enumerate(info["targets"]):
            if not isinstance(target, Mapping):
                raise InputError(
                    "drug target entries must be objects",
                    details={"ID": compound_id, "target_index": position},
                )
            evidence = _clean(target.get("evidence")).casefold()
            if evidence == "known":
                known.append((position, target))
            elif evidence == "predicted":
                rank_value = target.get("prediction_rank")
                if isinstance(rank_value, bool) or not isinstance(rank_value, (int, float)):
                    raise InputError(
                        "predicted targets require numeric prediction_rank",
                        details={"ID": compound_id, "target_index": position},
                    )
                rank = int(rank_value)
                if float(rank_value) != rank or not 1 <= rank <= 10:
                    raise InputError(
                        "prediction_rank must be an integer from 1 to 10",
                        details={"ID": compound_id, "target_index": position},
                    )
                predicted.append((rank, position, target))
            else:
                raise InputError(
                    "drug target evidence must be known or predicted",
                    details={"ID": compound_id, "target_index": position},
                )
        ordered = [item for _, item in known]
        ordered.extend(item for _, _, item in sorted(predicted, key=lambda value: value[:2]))
        selected: list[TargetRecord] = []
        seen_symbols: set[str] = set()
        for target in ordered:
            symbol = _symbol(target.get("gene_symbol"))
            if not symbol or symbol in seen_symbols:
                continue
            seen_symbols.add(symbol)
            evidence = _clean(target.get("evidence")).casefold()
            rank = "Known" if evidence == "known" else str(int(target["prediction_rank"]))
            score_value = target.get("score")
            score = "" if score_value is None else _clean(score_value)
            selected.append(
                TargetRecord(
                    symbol=symbol,
                    uniprot_id=_clean(target.get("uniprot_id")),
                    evidence=evidence,
                    score=score,
                    rank=rank,
                )
            )
            if len(selected) >= top_n:
                break
        output[compound_id.strip()] = selected
    return output


def _read_optional_table(path: Path | None, *, label: str) -> list[dict[str, str]]:
    if path is None:
        return []
    rows: list[dict[str, str]] = []
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter="\t")
            _require_columns(reader.fieldnames, ("input_type", "value"), label)
            for row_number, row in enumerate(reader, start=2):
                input_type = _clean(row.get("input_type"))
                value = _clean(row.get("value"))
                if not input_type and not value:
                    continue
                if not input_type or not value:
                    raise InputError(
                        f"{label} contains an incomplete record",
                        details={"row": row_number},
                    )
                rows.append({"input_type": input_type, "value": value})
    except InputError:
        raise
    except (OSError, UnicodeError, csv.Error) as exc:
        raise InputError(f"{label} could not be read", details={"path": str(path)}) from exc
    return rows


def _unique_node_by_name(
    base_nodes: Sequence[Mapping[str, str]], *, node_type: str, name: str, label: str
) -> str:
    matches = [
        row["node_id"]
        for row in base_nodes
        if row["node_type"] == node_type and row["node_name"].casefold() == name.casefold()
    ]
    if not matches:
        raise InputError(f"{label} did not match a base-graph node", details={"value": name})
    if len(matches) > 1:
        raise InputError(
            f"{label} matched multiple base-graph nodes; use an explicit node ID",
            details={"value": name, "node_ids": matches[:20]},
        )
    return matches[0]


def _atomic_write_table(
    path: Path,
    rows: Iterable[Mapping[str, object]],
    *,
    columns: Sequence[str],
    delimiter: str,
    allowed_root: Path,
) -> None:
    target = ensure_within(path, allowed_root)
    target.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        dir=target.parent, prefix=f".{target.name}.", suffix=".tmp"
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=list(columns),
                delimiter=delimiter,
                lineterminator="\n",
                extrasaction="ignore",
            )
            writer.writeheader()
            writer.writerows(rows)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, target)
    except BaseException:
        temporary.unlink(missing_ok=True)
        raise


def _temporary_writer(
    stack: ExitStack, target: Path, *, columns: Sequence[str], delimiter: str
) -> tuple[Path, csv.DictWriter]:
    target.parent.mkdir(parents=True, exist_ok=True)
    descriptor, name = tempfile.mkstemp(dir=target.parent, prefix=f".{target.name}.", suffix=".tmp")
    temporary = Path(name)
    raw_handle = os.fdopen(descriptor, "w", encoding="utf-8", newline="")
    handle: TextIO = stack.enter_context(raw_handle)
    writer = csv.DictWriter(
        handle,
        fieldnames=list(columns),
        delimiter=delimiter,
        lineterminator="\n",
        extrasaction="ignore",
    )
    writer.writeheader()
    return temporary, writer


def _active_bits(row: Any) -> list[int]:
    if hasattr(row, "indices"):
        return [int(value) for value in row.indices]
    if hasattr(row, "nonzero"):
        result = row.nonzero()
        values = result[-1] if isinstance(result, tuple) else result
        if hasattr(values, "tolist"):
            values = values.tolist()
        return [int(value) for value in values]
    return [index for index, value in enumerate(row) if value]


def _fingerprint_edges(
    compounds: Sequence[CompoundRecord],
    *,
    fingerprint: Any,
    chunk_size: int,
    invalid: list[dict[str, str]],
) -> Iterable[EdgeCandidate]:
    eligible = [record for record in compounds if record.is_new_node and record.canonical_smiles]
    invalid_ids = {row["ID"] for row in invalid}
    for start in range(0, len(eligible), chunk_size):
        chunk = eligible[start : start + chunk_size]
        try:
            matrix = fingerprint.transform([record.canonical_smiles for record in chunk])
            transformed = [(record, matrix[index]) for index, record in enumerate(chunk)]
        except Exception:
            transformed = []
            for record in chunk:
                try:
                    one = fingerprint.transform([record.canonical_smiles])
                    transformed.append((record, one[0]))
                except Exception:
                    if record.library_id not in invalid_ids:
                        invalid.append(
                            {
                                "ID": record.library_id,
                                "SMILES": record.smiles,
                                "reason": "pubchemfp_failed",
                            }
                        )
                        invalid_ids.add(record.library_id)
        for record, fingerprint_row in transformed:
            for bit in sorted(set(_active_bits(fingerprint_row))):
                yield EdgeCandidate(
                    x_id=record.node_id,
                    y_id=f"substructure:pubchemfp:{bit:03d}",
                    relation="drug_substructure",
                    relation_raw="has_substructure",
                    edge_type="small_molecule_substructure",
                    source_db="PubChemFP_scikit-fingerprints",
                    evidence="bit_1",
                    original_relation=f"PubchemFP{bit}",
                )


def _base_edges(path: Path) -> Iterable[EdgeCandidate]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            _require_columns(reader.fieldnames, EDGE_COLUMNS, "base edge table")
            for row in reader:
                yield EdgeCandidate(
                    x_id=_clean(row.get("x_id")),
                    y_id=_clean(row.get("y_id")),
                    relation=_clean(row.get("relation")),
                    relation_raw=_clean(row.get("relation_raw")),
                    edge_type=_clean(row.get("edge_type")),
                    source_db=_clean(row.get("source_db")),
                    evidence=_clean(row.get("evidence")),
                    score=_clean(row.get("score")),
                    rank=_clean(row.get("rank")),
                    original_relation=_clean(row.get("original_relation")),
                )
    except InputError:
        raise
    except (OSError, UnicodeError, csv.Error) as exc:
        raise ResourceError(
            "base edge table could not be read", details={"path": str(path)}
        ) from exc


def _write_edges(
    *,
    context: RunContext,
    edge_path: Path,
    kg_path: Path,
    candidates: Iterable[tuple[str, EdgeCandidate]],
    node_by_id: Mapping[str, Mapping[str, str]],
) -> dict[str, Any]:
    seen: set[tuple[str, str, str]] = set()
    relation_counts: Counter[str] = Counter()
    source_counts: Counter[str] = Counter()
    candidate_counts: Counter[str] = Counter()
    before_dedup = after_dedup = missing_endpoint_count = 0
    edge_temp: Path | None = None
    kg_temp: Path | None = None
    try:
        with ExitStack() as stack:
            edge_temp, edge_writer = _temporary_writer(
                stack, edge_path, columns=EDGE_COLUMNS, delimiter=","
            )
            kg_temp, kg_writer = _temporary_writer(
                stack, kg_path, columns=KG_COLUMNS, delimiter=","
            )
            for category, candidate in candidates:
                candidate_counts[category] += 1
                if not candidate.x_id or not candidate.y_id or candidate.x_id == candidate.y_id:
                    continue
                if candidate.x_id not in node_by_id or candidate.y_id not in node_by_id:
                    missing_endpoint_count += 1
                    continue
                before_dedup += 1
                key = (candidate.x_id, candidate.relation, candidate.y_id)
                if key in seen:
                    continue
                seen.add(key)
                after_dedup += 1
                edge_id = f"edge:{after_dedup:09d}"
                source = node_by_id[candidate.x_id]
                target = node_by_id[candidate.y_id]
                edge_row = {
                    "edge_id": edge_id,
                    "x_index": source["node_index"],
                    "y_index": target["node_index"],
                    "x_id": candidate.x_id,
                    "y_id": candidate.y_id,
                    "x_type": source["node_type"],
                    "y_type": target["node_type"],
                    "relation": candidate.relation,
                    "relation_raw": candidate.relation_raw,
                    "edge_type": candidate.edge_type,
                    "source_db": candidate.source_db,
                    "evidence": candidate.evidence,
                    "score": candidate.score,
                    "rank": candidate.rank,
                    "original_relation": candidate.original_relation,
                }
                edge_writer.writerow(edge_row)
                kg_writer.writerow(
                    {
                        "x_type": source["node_type"],
                        "x_id": candidate.x_id,
                        "x_index": source["node_index"],
                        "x_name": source["node_name"],
                        "x_source": source["node_source"],
                        "relation": candidate.relation,
                        "y_type": target["node_type"],
                        "y_id": candidate.y_id,
                        "y_index": target["node_index"],
                        "y_name": target["node_name"],
                        "y_source": target["node_source"],
                        "source_db": candidate.source_db,
                        "evidence": candidate.evidence,
                        "score": candidate.score,
                        "rank": candidate.rank,
                        "original_relation": candidate.original_relation,
                        "edge_type": candidate.edge_type,
                        "edge_id": edge_id,
                    }
                )
                relation_counts[candidate.relation] += 1
                source_counts[candidate.source_db] += 1
        os.replace(edge_temp, ensure_within(edge_path, context.output_dir))
        edge_temp = None
        os.replace(kg_temp, ensure_within(kg_path, context.output_dir))
        kg_temp = None
    finally:
        if edge_temp is not None:
            edge_temp.unlink(missing_ok=True)
        if kg_temp is not None:
            kg_temp.unlink(missing_ok=True)
    return {
        "num_edges_before_dedup": before_dedup,
        "num_edges_after_dedup": after_dedup,
        "num_edges_deduplicated": before_dedup - after_dedup,
        "num_edges_with_missing_endpoints": missing_endpoint_count,
        "relation_counts": dict(sorted(relation_counts.items())),
        "source_db_counts": dict(sorted(source_counts.items())),
        "edge_candidate_counts": dict(sorted(candidate_counts.items())),
    }


def _validate_settings(settings: KGConstructionConfig, disease: DiseaseConfig) -> None:
    if disease.species != "human":
        raise InputError(
            f"unsupported disease species {disease.species!r}; KG resources support human only"
        )
    if settings.netinfer_dti_top_n < 1:
        raise InputError("kg.construction.netinfer_dti_top_n must be positive")
    if settings.netinfer_dti_top_n > 10:
        raise InputError("kg.construction.netinfer_dti_top_n cannot exceed NetInfer top 10")
    if settings.netinfer_dti_selection != "known_first_then_predicted_rank":
        raise InputError("KG NetInfer DTI selection must remain known_first_then_predicted_rank")
    if settings.pubchem_fingerprint_chunk_size < 1:
        raise InputError("PubChemFP chunk size must be positive")
    for label, value in (
        ("disease.name", disease.name),
        ("disease.slug", disease.slug),
        ("disease.custom_node_id", disease.custom_node_id),
        ("disease.source_tag", disease.source_tag),
        ("kg.construction.candidate_source_tag", settings.candidate_source_tag),
        ("kg.construction.user_drug_node_prefix", settings.user_drug_node_prefix),
    ):
        if not _clean(value):
            raise InputError(f"{label} must be non-empty")


def _operation(
    execution: NodeExecution,
    *,
    context: RunContext,
    compounds_path: str | Path,
    disease_genes_path: str | Path,
    drug_targets_path: str | Path | None,
    drug_targets_manifest_path: str | Path | None,
    base_nodes_path: str | Path,
    base_edges_path: str | Path,
    base_manifest_path: str | Path,
    base_drug_smiles_path: str | Path,
    target_mapping_path: str | Path,
    positive_drugs_path: str | Path | None,
    disease_links_path: str | Path | None,
    settings: KGConstructionConfig,
    disease: DiseaseConfig,
    fingerprint_factory: Callable[[int], Any] | None,
) -> None:
    started = time.perf_counter()
    _validate_settings(settings, disease)
    chem, standardizer = _load_chemistry()
    fingerprint = (
        _load_fingerprint(workers=settings.pubchem_fingerprint_workers)
        if fingerprint_factory is None
        else fingerprint_factory(settings.pubchem_fingerprint_workers)
    )
    if fingerprint is None or not callable(getattr(fingerprint, "transform", None)):
        raise EnvironmentError(
            "PubChemFP dependency must provide a callable transform method",
            retryable=False,
        )

    compounds_file = _resolve_run_file(context, compounds_path, label="normalized compounds")
    disease_genes_file = _resolve_run_file(
        context, disease_genes_path, label="normalized disease genes"
    )
    positive_file = _resolve_optional_run_file(
        context, positive_drugs_path, label="positive_drugs.tsv"
    )
    disease_links_file = _resolve_optional_run_file(
        context, disease_links_path, label="disease_links.tsv"
    )
    drug_targets_file, drug_targets_artifact = _resolve_drug_targets(
        context,
        drug_targets_path=drug_targets_path,
        drug_targets_manifest_path=drug_targets_manifest_path,
    )
    if drug_targets_artifact is not None:
        execution.input_artifact_ids = tuple(
            dict.fromkeys((*execution.input_artifact_ids, drug_targets_artifact.artifact_id))
        )

    base_nodes_file = _resolve_resource_file(context, base_nodes_path, label="base node table")
    base_edges_file = _resolve_resource_file(context, base_edges_path, label="base edge table")
    base_manifest_file = _resolve_resource_file(
        context, base_manifest_path, label="base graph manifest"
    )
    base_smiles_file = _resolve_resource_file(
        context, base_drug_smiles_path, label="base drug SMILES table"
    )
    try:
        target_mapping_file = _resolve_run_file(
            context, target_mapping_path, label="NetInfer target mapping"
        )
        mapping_is_resource = False
    except InputError:
        target_mapping_file = _resolve_resource_file(
            context, target_mapping_path, label="NetInfer target mapping"
        )
        mapping_is_resource = True
    target_mapping_hash = sha256_file(target_mapping_file)
    resource_hashes = {
        "resources.kg.node_table": sha256_file(base_nodes_file),
        "resources.kg.edge_table": sha256_file(base_edges_file),
        "resources.kg.manifest": sha256_file(base_manifest_file),
        "resources.kg.drug_smiles": sha256_file(base_smiles_file),
    }
    if mapping_is_resource:
        resource_hashes["resources.netinfer.target_mapping"] = target_mapping_hash
    execution.resource_hashes.update(resource_hashes)

    try:
        base_manifest = json.loads(base_manifest_file.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise ResourceError("base graph manifest could not be read") from exc
    if not isinstance(base_manifest, dict):
        raise ResourceError("base graph manifest must contain a JSON object")

    base_nodes = _read_nodes(base_nodes_file)
    base_node_by_id = {row["node_id"]: row for row in base_nodes}
    if disease.custom_node_id in base_node_by_id:
        raise InputError(
            "configured custom disease node already exists in the base graph",
            details={"custom_node_id": disease.custom_node_id},
        )
    base_drug_ids = {row["node_id"] for row in base_nodes if row["node_type"] == "drug"}
    base_disease_ids = {row["node_id"] for row in base_nodes if row["node_type"] == "disease"}

    compounds, invalid_smiles = _read_compounds(
        compounds_file, chem=chem, standardizer=standardizer
    )
    disease_genes = _read_disease_genes(disease_genes_file)
    target_mapping = _read_target_mapping(target_mapping_file)
    targets = _read_drug_targets(drug_targets_file, top_n=settings.netinfer_dti_top_n)
    positive_inputs = _read_optional_table(positive_file, label="positive_drugs.tsv")
    disease_link_inputs = _read_optional_table(disease_links_file, label="disease_links.tsv")

    canonical_to_nodes: dict[str, set[str]] = defaultdict(set)
    invalid_base_smiles = 0
    base_smiles_rows = 0
    try:
        with base_smiles_file.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter="\t")
            _require_columns(
                reader.fieldnames,
                ("node_id", "node_name", "smiles"),
                "base drug SMILES table",
            )
            for row in reader:
                node_id = _clean(row.get("node_id"))
                if node_id not in base_drug_ids:
                    continue
                base_smiles_rows += 1
                canonical = _standardize_smiles(
                    row.get("smiles"), chem=chem, standardizer=standardizer
                )
                if not canonical:
                    invalid_base_smiles += 1
                    continue
                canonical_to_nodes[canonical].add(node_id)
    except InputError:
        raise
    except (OSError, UnicodeError, csv.Error) as exc:
        raise ResourceError("base drug SMILES table could not be read") from exc
    unique_base_smiles = {
        canonical: next(iter(node_ids))
        for canonical, node_ids in canonical_to_nodes.items()
        if len(node_ids) == 1
    }
    ambiguous_base_smiles = {
        canonical for canonical, node_ids in canonical_to_nodes.items() if len(node_ids) > 1
    }

    existing_node_ids = set(base_node_by_id)
    user_source = settings.candidate_source_tag.rstrip(":") or settings.candidate_source_tag
    next_user_node = 1

    def new_user_node_id() -> str:
        nonlocal next_user_node
        while True:
            node_id = f"{settings.user_drug_node_prefix}{next_user_node:08d}"
            next_user_node += 1
            if node_id not in existing_node_ids:
                existing_node_ids.add(node_id)
                return node_id

    match_rows: list[dict[str, str]] = []
    new_nodes: list[dict[str, str]] = []
    library_to_node: dict[str, str] = {}
    for compound in compounds:
        matched = ""
        if compound.canonical_smiles in unique_base_smiles:
            matched = unique_base_smiles[compound.canonical_smiles]
            compound.node_id = matched
            compound.is_new_node = False
            compound.match_type = "base_canonical_smiles"
            compound.matched_base_node_id = matched
            base_node_by_id[matched]["source_ids"] = _append_source_ids(
                base_node_by_id[matched]["source_ids"],
                f"{settings.candidate_source_tag}{compound.library_id}",
            )
        else:
            compound.node_id = new_user_node_id()
            if compound.canonical_smiles in ambiguous_base_smiles:
                compound.match_type = "ambiguous_base_canonical_smiles_kept_new"
            elif not compound.canonical_smiles:
                compound.match_type = "invalid_smiles_kept_new"
            description = ["UserLibrary"]
            if compound.formula:
                description.append(f"Formula={compound.formula}")
            if compound.molwt:
                description.append(f"MolWt={compound.molwt}")
            new_nodes.append(
                {
                    "node_index": "",
                    "node_id": compound.node_id,
                    "node_type": "drug",
                    "node_name": compound.name,
                    "node_source": user_source,
                    "source_ids": _append_source_ids(
                        f"{settings.candidate_source_tag}{compound.library_id}",
                        f"CAS:{compound.cas}" if compound.cas else "",
                    ),
                    "description": ";".join(description),
                    "node_type_raw": "small_molecule",
                }
            )
        library_to_node[compound.library_id] = compound.node_id
        match_rows.append(
            {
                "library_id": compound.library_id,
                "input_smiles": compound.smiles,
                "canonical_smiles": compound.canonical_smiles,
                "output_node_id": compound.node_id,
                "is_new_node": "1" if compound.is_new_node else "0",
                "match_type": compound.match_type,
                "matched_base_node_id": matched,
                "matched_base_node_name": base_node_by_id.get(matched, {}).get("node_name", ""),
            }
        )

    unmapped_occurrences: Counter[str] = Counter()
    unmapped_compounds: dict[str, set[str]] = defaultdict(set)

    def ensure_protein(entrez_id: str, symbol: str, source: str) -> str:
        node_id = _protein_id(entrez_id)
        if not node_id:
            return ""
        existing = base_node_by_id.get(node_id)
        if existing is not None and existing["node_type"] != "gene/protein":
            raise ResourceError(
                "protein node ID collides with a non-protein base node",
                details={"node_id": node_id, "node_type": existing["node_type"]},
            )
        if node_id not in existing_node_ids:
            existing_node_ids.add(node_id)
            new_nodes.append(
                {
                    "node_index": "",
                    "node_id": node_id,
                    "node_type": "gene/protein",
                    "node_name": symbol or entrez_id,
                    "node_source": "NCBI",
                    "source_ids": f"NCBI:{entrez_id}",
                    "description": source,
                    "node_type_raw": "protein",
                }
            )
        return node_id

    for library_id, records in targets.items():
        if library_id not in library_to_node:
            continue
        for target in records:
            entrez_id = target_mapping.get(target.symbol, "")
            if entrez_id:
                ensure_protein(entrez_id, target.symbol, "NetInfer")
            else:
                unmapped_occurrences[target.symbol] += 1
                unmapped_compounds[target.symbol].add(library_id)
    for gene_symbol, entrez_id in disease_genes:
        ensure_protein(entrez_id, gene_symbol, "DiseaseGeneSet")

    new_nodes.append(
        {
            "node_index": "",
            "node_id": disease.custom_node_id,
            "node_type": "disease",
            "node_name": disease.name,
            "node_source": disease.source_tag,
            "source_ids": _append_source_ids(
                disease.identifier, f"{disease.source_tag}:{disease.slug}"
            ),
            "description": disease.description or disease.name,
            "node_type_raw": "disease",
        }
    )
    all_nodes = [*base_nodes, *new_nodes]
    for index, row in enumerate(all_nodes):
        row["node_index"] = str(index)
    node_by_id = {row["node_id"]: row for row in all_nodes}
    if len(node_by_id) != len(all_nodes):
        raise OutputContractError("constructed node table contains duplicate node IDs")

    resolved_positive: list[dict[str, str]] = []
    for row in positive_inputs:
        input_type, value = row["input_type"], row["value"]
        if input_type == "library_id":
            if value not in library_to_node:
                raise InputError(
                    "positive library_id is absent from normalized compounds",
                    details={"value": value},
                )
            node_id = library_to_node[value]
        elif input_type == "base_drug_id":
            if value not in base_drug_ids:
                raise InputError(
                    "positive base_drug_id is absent from the base graph",
                    details={"value": value},
                )
            node_id = value
        elif input_type == "base_drug_name":
            node_id = _unique_node_by_name(
                base_nodes, node_type="drug", name=value, label="base_drug_name"
            )
        else:
            raise InputError(
                "positive_drugs.tsv contains an unsupported input_type",
                details={"input_type": input_type},
            )
        resolved_positive.append({"input_type": input_type, "value": value, "node_id": node_id})

    resolved_links: list[dict[str, str]] = []
    for row in disease_link_inputs:
        input_type, value = row["input_type"], row["value"]
        if input_type == "base_disease_id":
            if value not in base_disease_ids:
                raise InputError(
                    "linked base_disease_id is absent from the base graph",
                    details={"value": value},
                )
            node_id = value
        elif input_type == "base_disease_name":
            node_id = _unique_node_by_name(
                base_nodes,
                node_type="disease",
                name=value,
                label="base_disease_name",
            )
        else:
            raise InputError(
                "disease_links.tsv contains an unsupported input_type",
                details={"input_type": input_type},
            )
        resolved_links.append({"input_type": input_type, "value": value, "node_id": node_id})

    output_paths = {
        artifact_type: _resolve_output(context, filename)
        for artifact_type, filename in OUTPUT_FILES.items()
    }
    _atomic_write_table(
        output_paths["kg_nodes"],
        all_nodes,
        columns=NODE_COLUMNS,
        delimiter="\t",
        allowed_root=context.output_dir,
    )
    _atomic_write_table(
        output_paths["kg_drug_match_report"],
        match_rows,
        columns=MATCH_REPORT_COLUMNS,
        delimiter="\t",
        allowed_root=context.output_dir,
    )

    def candidates() -> Iterable[tuple[str, EdgeCandidate]]:
        for edge in _base_edges(base_edges_file):
            yield "base", edge
        for library_id, drug_node in library_to_node.items():
            for target in targets.get(library_id, ()):
                entrez_id = target_mapping.get(target.symbol, "")
                protein = _protein_id(entrez_id)
                if not protein:
                    continue
                yield (
                    "netinfer_dti",
                    EdgeCandidate(
                        x_id=drug_node,
                        y_id=protein,
                        relation="drug_protein",
                        relation_raw="target",
                        edge_type="small_molecule_protein",
                        source_db="NetInfer",
                        evidence=target.evidence,
                        score=target.score,
                        rank=target.rank,
                        original_relation=(
                            f"NetInfer_top{settings.netinfer_dti_top_n}_known_first"
                        ),
                    ),
                )
        for edge in _fingerprint_edges(
            compounds,
            fingerprint=fingerprint,
            chunk_size=settings.pubchem_fingerprint_chunk_size,
            invalid=invalid_smiles,
        ):
            yield "pubchemfp", edge
        for gene_symbol, entrez_id in disease_genes:
            del gene_symbol
            yield (
                "disease_gene",
                EdgeCandidate(
                    x_id=_protein_id(entrez_id),
                    y_id=disease.custom_node_id,
                    relation="disease_protein",
                    relation_raw="associated_with",
                    edge_type="protein_disease",
                    source_db="disease_gene_file",
                    evidence="custom_gene_set",
                    original_relation="disease_gene_set",
                ),
            )
        for row in resolved_positive:
            yield (
                "positive_drug",
                EdgeCandidate(
                    x_id=row["node_id"],
                    y_id=disease.custom_node_id,
                    relation="drug_disease",
                    relation_raw="drug_disease",
                    edge_type="small_molecule_disease",
                    source_db="custom_positive",
                    evidence="custom_positive_drug",
                    original_relation="drug_disease",
                ),
            )
        for row in resolved_links:
            yield (
                "disease_link",
                EdgeCandidate(
                    x_id=disease.custom_node_id,
                    y_id=row["node_id"],
                    relation="disease_disease",
                    relation_raw="parent-child",
                    edge_type="disease_disease",
                    source_db="custom_disease_link",
                    evidence="custom_disease_link",
                    original_relation="parent-child",
                ),
            )

    edge_metrics = _write_edges(
        context=context,
        edge_path=output_paths["kg_edges"],
        kg_path=output_paths["kg_graph"],
        candidates=candidates(),
        node_by_id=node_by_id,
    )
    invalid_smiles.sort(key=lambda row: (row["ID"], row["reason"]))
    _atomic_write_table(
        output_paths["kg_invalid_smiles"],
        invalid_smiles,
        columns=INVALID_SMILES_COLUMNS,
        delimiter="\t",
        allowed_root=context.output_dir,
    )
    unmapped_rows = [
        {
            "gene_symbol": symbol,
            "occurrence_count": str(unmapped_occurrences[symbol]),
            "compound_ids": ";".join(sorted(unmapped_compounds[symbol])),
            "reason": "missing_symbol_to_entrez_mapping",
        }
        for symbol in sorted(unmapped_occurrences)
    ]
    _atomic_write_table(
        output_paths["kg_unmapped_targets"],
        unmapped_rows,
        columns=UNMAPPED_TARGET_COLUMNS,
        delimiter="\t",
        allowed_root=context.output_dir,
    )

    drugs_without_targets = sum(1 for library_id in library_to_node if not targets.get(library_id))
    node_type_counts = Counter(row["node_type"] for row in all_nodes)
    metrics: dict[str, Any] = {
        "num_base_nodes": len(base_nodes),
        "num_base_edges": edge_metrics["edge_candidate_counts"].get("base", 0),
        "num_user_library_compounds": len(compounds),
        "num_user_drugs_matched_to_base_by_smiles": sum(
            not record.is_new_node for record in compounds
        ),
        "num_new_user_drug_nodes": sum(record.is_new_node for record in compounds),
        "num_invalid_user_smiles_for_matching": sum(
            row["reason"] == "invalid_smiles" for row in invalid_smiles
        ),
        "num_invalid_smiles_for_pubchemfp": len(invalid_smiles),
        "num_ambiguous_base_canonical_smiles": len(ambiguous_base_smiles),
        "num_invalid_base_smiles": invalid_base_smiles,
        "num_base_drug_smiles_rows": base_smiles_rows,
        "num_new_protein_nodes": sum(row["node_type"] == "gene/protein" for row in new_nodes),
        "num_drugs_without_netinfer_records": drugs_without_targets,
        "num_unmapped_target_symbols": len(unmapped_rows),
        "num_netinfer_dti_edges_added": edge_metrics["edge_candidate_counts"].get(
            "netinfer_dti", 0
        ),
        "num_user_substructure_edges_added": edge_metrics["edge_candidate_counts"].get(
            "pubchemfp", 0
        ),
        "num_disease_gene_edges_added": edge_metrics["edge_candidate_counts"].get(
            "disease_gene", 0
        ),
        "num_positive_drug_edges_added": edge_metrics["edge_candidate_counts"].get(
            "positive_drug", 0
        ),
        "num_disease_link_edges_added": edge_metrics["edge_candidate_counts"].get(
            "disease_link", 0
        ),
        "num_nodes": len(all_nodes),
        "node_type_counts": dict(sorted(node_type_counts.items())),
        **edge_metrics,
    }
    input_records: dict[str, Any] = {
        "compounds": {
            "path": context.relative_path(compounds_file),
            "sha256": sha256_file(compounds_file),
        },
        "disease_genes": {
            "path": context.relative_path(disease_genes_file),
            "sha256": sha256_file(disease_genes_file),
        },
        "drug_targets": {
            "path": context.relative_path(drug_targets_file),
            "sha256": sha256_file(drug_targets_file),
            "artifact_id": None
            if drug_targets_artifact is None
            else drug_targets_artifact.artifact_id,
        },
        "positive_drugs": None
        if positive_file is None
        else {
            "path": context.relative_path(positive_file),
            "sha256": sha256_file(positive_file),
        },
        "disease_links": None
        if disease_links_file is None
        else {
            "path": context.relative_path(disease_links_file),
            "sha256": sha256_file(disease_links_file),
        },
    }
    base_resource_id = _clean(base_manifest.get("resource_id")) or (
        f"kg-base:{_clean(base_manifest.get('variant')) or 'unknown'}"
    )
    construction_manifest = {
        "schema_version": SCHEMA_VERSION,
        "variant": "run_base_graph_plus_user_inputs",
        "disease": {
            "name": disease.name,
            "slug": disease.slug,
            "identifier": disease.identifier,
            "custom_node_id": disease.custom_node_id,
            "species": disease.species,
            "tissue": disease.tissue,
            "description": disease.description,
            "source_tag": disease.source_tag,
        },
        "configuration": {
            "candidate_source_tag": settings.candidate_source_tag,
            "user_drug_node_prefix": settings.user_drug_node_prefix,
            "netinfer_dti_top_n": settings.netinfer_dti_top_n,
            "netinfer_dti_selection": settings.netinfer_dti_selection,
            "pubchem_fingerprint": "PubChemFP",
            "pubchem_fingerprint_workers": settings.pubchem_fingerprint_workers,
            "pubchem_fingerprint_chunk_size": settings.pubchem_fingerprint_chunk_size,
        },
        "inputs": input_records,
        "resources": {
            "base_graph": {
                "resource_id": base_resource_id,
                "variant": _clean(base_manifest.get("variant")),
                "manifest_sha256": resource_hashes["resources.kg.manifest"],
                "node_table_sha256": resource_hashes["resources.kg.node_table"],
                "edge_table_sha256": resource_hashes["resources.kg.edge_table"],
                "drug_smiles_sha256": resource_hashes["resources.kg.drug_smiles"],
            },
            "target_mapping": {
                "source": "fixed_resource" if mapping_is_resource else "run_artifact",
                "sha256": target_mapping_hash,
            },
        },
        "optional_inputs": {
            "positive_drugs": resolved_positive,
            "disease_links": resolved_links,
        },
        "metrics": metrics,
        "software": {
            "lipid_screening_agent": __version__,
            "rdkit": _package_version("rdkit"),
            "scikit_fingerprints": _package_version("scikit-fingerprints"),
        },
        "files": {
            key: context.relative_path(path)
            for key, path in output_paths.items()
            if key != "kg_construction_manifest"
        },
        "file_sha256": {
            key: sha256_file(path)
            for key, path in output_paths.items()
            if key != "kg_construction_manifest"
        },
    }
    atomic_write_json(
        output_paths["kg_construction_manifest"],
        construction_manifest,
        allowed_root=context.output_dir,
    )
    execution.update_metrics(metrics)
    if invalid_smiles:
        execution.warn(
            f"{len(invalid_smiles)} compound record(s) lacked a usable PubChemFP fingerprint"
        )
    if unmapped_rows:
        execution.warn(
            f"{len(unmapped_rows)} NetInfer target symbol(s) could not be mapped to Entrez"
        )
    if edge_metrics["num_edges_with_missing_endpoints"]:
        missing_endpoint_count = edge_metrics["num_edges_with_missing_endpoints"]
        execution.warn(f"Skipped {missing_endpoint_count} edge(s) with missing endpoints")
    for artifact_type, path in output_paths.items():
        execution.add_output(artifact_type, path)
    execution.metric("elapsed_seconds", time.perf_counter() - started)


def kg_construct_graph(
    *,
    context: RunContext,
    compounds_path: str | Path,
    disease_genes_path: str | Path,
    base_nodes_path: str | Path,
    base_edges_path: str | Path,
    base_manifest_path: str | Path,
    base_drug_smiles_path: str | Path,
    target_mapping_path: str | Path,
    settings: KGConstructionConfig,
    disease: DiseaseConfig,
    config_hash: str,
    code_version: str,
    drug_targets_path: str | Path | None = None,
    drug_targets_manifest_path: str | Path | None = None,
    positive_drugs_path: str | Path | None = None,
    disease_links_path: str | Path | None = None,
    task_id: str = "main",
    attempt: int = 1,
    input_artifact_ids: Sequence[str] = (),
    fingerprint_factory: Callable[[int], Any] | None = None,
) -> NodeResult:
    """Construct all Stage 06 graph artifacts without mutating base resources."""

    return execute_node(
        lambda execution: _operation(
            execution,
            context=context,
            compounds_path=compounds_path,
            disease_genes_path=disease_genes_path,
            drug_targets_path=drug_targets_path,
            drug_targets_manifest_path=drug_targets_manifest_path,
            base_nodes_path=base_nodes_path,
            base_edges_path=base_edges_path,
            base_manifest_path=base_manifest_path,
            base_drug_smiles_path=base_drug_smiles_path,
            target_mapping_path=target_mapping_path,
            positive_drugs_path=positive_drugs_path,
            disease_links_path=disease_links_path,
            settings=settings,
            disease=disease,
            fingerprint_factory=fingerprint_factory,
        ),
        context=context,
        node_id=NODE_ID,
        task_id=task_id,
        attempt=attempt,
        config_hash=config_hash,
        code_version=code_version,
        input_artifact_ids=input_artifact_ids,
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Construct a run-local KG from a configured immutable base graph."
    )
    add_common_runner_arguments(parser)
    parser.add_argument("--compounds", required=True, type=Path)
    parser.add_argument("--disease-genes", required=True, type=Path)
    target_group = parser.add_mutually_exclusive_group(required=True)
    target_group.add_argument("--drug-targets", type=Path)
    target_group.add_argument("--drug-targets-manifest", type=Path)
    parser.add_argument("--base-nodes", required=True, type=Path)
    parser.add_argument("--base-edges", required=True, type=Path)
    parser.add_argument("--base-manifest", required=True, type=Path)
    parser.add_argument("--base-drug-smiles", required=True, type=Path)
    parser.add_argument("--target-mapping", required=True, type=Path)
    parser.add_argument("--positive-drugs", type=Path)
    parser.add_argument("--disease-links", type=Path)
    add_execution_identity_arguments(parser)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    namespace = build_parser().parse_args(argv)
    arguments = CommonRunnerArguments.from_namespace(namespace)
    project_root = Path(__file__).resolve().parents[4]
    environment = load_common_runner_environment(arguments, project_root=project_root)
    task_id, attempt, input_artifact_ids = execution_identity(namespace)
    result = kg_construct_graph(
        context=environment.context,
        compounds_path=namespace.compounds,
        disease_genes_path=namespace.disease_genes,
        drug_targets_path=namespace.drug_targets,
        drug_targets_manifest_path=namespace.drug_targets_manifest,
        base_nodes_path=namespace.base_nodes,
        base_edges_path=namespace.base_edges,
        base_manifest_path=namespace.base_manifest,
        base_drug_smiles_path=namespace.base_drug_smiles,
        target_mapping_path=namespace.target_mapping,
        positive_drugs_path=namespace.positive_drugs,
        disease_links_path=namespace.disease_links,
        settings=environment.config.kg.construction,
        disease=environment.config.disease,
        config_hash=environment.config_hash,
        code_version=__version__,
        task_id=task_id,
        attempt=attempt,
        input_artifact_ids=input_artifact_ids,
    )
    sys.stdout.write(result.to_json() + "\n")
    return 0 if result.status is NodeStatus.SUCCEEDED else 1


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())


__all__ = ["build_parser", "kg_construct_graph", "main"]
