"""Prepare the PPI LCC, mapped targets, and content-addressed disease distances."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import sys
import time
from collections import defaultdict
from collections.abc import Sequence
from pathlib import Path
from typing import Any

import networkx as nx
import numpy as np

from lipid_screening_agent import __version__
from lipid_screening_agent.artifacts import NodeResult, NodeStatus
from lipid_screening_agent.cli import (
    CommonRunnerArguments,
    add_common_runner_arguments,
    load_common_runner_environment,
)
from lipid_screening_agent.config.models import ProximityConfig
from lipid_screening_agent.runtime import (
    InputError,
    ResourceError,
    RunContext,
    atomic_write_json,
    hash_json,
    sha256_file,
)
from lipid_screening_agent.runtime.execution import NodeExecution, execute_node

from ..input_prepare._common import add_execution_identity_arguments, execution_identity
from ._common import (
    atomic_write_delimited,
    atomic_write_npz,
    resolve_cache_directories,
    resolve_proximity_output,
    resolve_resource_file,
    resolve_run_input_file,
    resolve_successful_drug_targets,
)
from .algorithms import ALGORITHM_VERSION, CACHE_FORMAT_VERSION, build_disease_distance_matrix

NODE_ID = "proximity_prepare_network"
NETWORK_MANIFEST_PATH = "network_manifest.json"
PREPARED_NETWORK_PATH = "prepared_network.npz"
PREPARED_TARGETS_PATH = "prepared_drug_targets.tsv"
SKIPPED_COMPOUNDS_PATH = "skipped_compounds.tsv"
UNMAPPED_TARGETS_PATH = "unmapped_targets.tsv"
PREPARATION_SCHEMA_VERSION = "1.0"


def _clean_gene_id(value: object) -> str | None:
    if value is None:
        return None
    cleaned = str(value).split(";", 1)[0].strip()
    return cleaned or None


def _symbol_key(value: object) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned.upper() if cleaned else None


def _read_ppi(path: Path) -> tuple[nx.Graph[str], int]:
    graph = nx.Graph()
    seen_edges: set[tuple[str, str]] = set()
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle, delimiter="\t")
            header = next(reader, None)
            if header is None or len(header) < 2:
                raise ResourceError("PPI table must contain at least two columns")
            for row_number, row in enumerate(reader, start=2):
                if len(row) < 2:
                    raise ResourceError(
                        "PPI row contains fewer than two columns",
                        details={"row": row_number},
                    )
                left, right = row[0].strip(), row[1].strip()
                if not left or not right or left.casefold() == "nan" or right.casefold() == "nan":
                    continue
                if left == right:
                    continue
                edge = (left, right) if left < right else (right, left)
                if edge in seen_edges:
                    continue
                seen_edges.add(edge)
                graph.add_edge(left, right)
    except ResourceError:
        raise
    except (OSError, UnicodeError, csv.Error) as exc:
        raise ResourceError("PPI table could not be read", details={"path": str(path)}) from exc
    if graph.number_of_nodes() == 0:
        raise InputError("PPI network is empty after removing invalid/self edges")
    return graph, len(seen_edges)


def _read_disease_genes(path: Path) -> tuple[list[str], int, int]:
    genes: list[str] = []
    seen: set[str] = set()
    input_count = duplicate_count = 0
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter="\t")
            if reader.fieldnames is None or not {"symbol", "entrez_id"}.issubset(reader.fieldnames):
                raise InputError("normalized disease genes must contain symbol and entrez_id")
            for row in reader:
                input_count += 1
                entrez_id = _clean_gene_id(row.get("entrez_id"))
                if entrez_id is None:
                    continue
                if entrez_id in seen:
                    duplicate_count += 1
                    continue
                seen.add(entrez_id)
                genes.append(entrez_id)
    except InputError:
        raise
    except (OSError, UnicodeError, csv.Error) as exc:
        raise InputError(
            "normalized disease genes could not be read",
            details={"path": str(path)},
        ) from exc
    if not genes:
        raise InputError("normalized disease gene input is empty")
    return genes, input_count, duplicate_count


def _mapping_from_rows(rows: Sequence[Sequence[object]], *, source_path: Path) -> dict[str, str]:
    if not rows:
        raise ResourceError("target mapping contains no rows")
    header = [str(value or "").strip() for value in rows[0]]
    aliases = {
        "symbol": {"Gene symbol", "gene_symbol", "symbol"},
        "entrez": {"Gene ID", "entrez_id", "gene_id"},
    }
    symbol_indices = [index for index, value in enumerate(header) if value in aliases["symbol"]]
    entrez_indices = [index for index, value in enumerate(header) if value in aliases["entrez"]]
    if len(symbol_indices) != 1 or len(entrez_indices) != 1:
        raise ResourceError(
            "target mapping must contain one gene-symbol and one Entrez-ID column",
            details={"path": str(source_path), "columns": header},
        )
    symbol_index, entrez_index = symbol_indices[0], entrez_indices[0]
    result: dict[str, str] = {}
    ambiguous: dict[str, set[str]] = defaultdict(set)
    for row in rows[1:]:
        if len(row) <= max(symbol_index, entrez_index):
            continue
        symbol = _symbol_key(row[symbol_index])
        entrez_id = _clean_gene_id(row[entrez_index])
        if symbol is None or entrez_id is None:
            continue
        previous = result.get(symbol)
        if previous is not None and previous != entrez_id:
            ambiguous[symbol].update((previous, entrez_id))
        else:
            result[symbol] = entrez_id
    if ambiguous:
        raise ResourceError(
            "target mapping contains symbols assigned to multiple Entrez IDs",
            details={
                "ambiguous_symbols": {
                    symbol: sorted(values)
                    for symbol, values in list(sorted(ambiguous.items()))[:20]
                }
            },
        )
    if not result:
        raise ResourceError("target mapping contains no usable symbol-to-Entrez records")
    return result


def _read_target_mapping(path: Path) -> dict[str, str]:
    if path.suffix.casefold() in {".xlsx", ".xlsm"}:
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover - depends on installed extras
            from lipid_screening_agent.runtime import EnvironmentError

            raise EnvironmentError(
                "openpyxl is required to read the NetInfer target workbook",
                retryable=False,
            ) from exc
        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
            if "Target information" not in workbook.sheetnames:
                raise ResourceError("target workbook is missing the 'Target information' sheet")
            sheet = workbook["Target information"]
            rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
            workbook.close()
        except ResourceError:
            raise
        except Exception as exc:
            raise ResourceError(
                "target workbook could not be read", details={"path": str(path)}
            ) from exc
        return _mapping_from_rows(rows, source_path=path)

    delimiter = "," if path.suffix.casefold() == ".csv" else "\t"
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = [tuple(row) for row in csv.reader(handle, delimiter=delimiter)]
    except (OSError, UnicodeError, csv.Error) as exc:
        raise ResourceError(
            "target mapping table could not be read", details={"path": str(path)}
        ) from exc
    return _mapping_from_rows(rows, source_path=path)


def _read_drug_targets(path: Path) -> dict[str, dict[str, Any]]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise InputError("NetInfer drug_targets.json could not be read") from exc
    if not isinstance(payload, dict):
        raise InputError("NetInfer drug_targets.json must be keyed by compound ID")
    normalized: dict[str, dict[str, Any]] = {}
    for compound_id, info in payload.items():
        if not isinstance(compound_id, str) or not compound_id.strip():
            raise InputError("drug_targets.json contains an empty compound ID")
        if not isinstance(info, dict) or not isinstance(info.get("targets"), list):
            raise InputError(
                "each drug_targets.json value must contain a targets list",
                details={"ID": compound_id},
            )
        targets: list[dict[str, Any]] = []
        predicted_rank = 0
        predicted_started = False
        seen_symbols: set[str] = set()
        for target_index, target in enumerate(info["targets"]):
            if not isinstance(target, dict):
                raise InputError(
                    "drug target entries must be objects",
                    details={"ID": compound_id, "target_index": target_index},
                )
            symbol = _symbol_key(target.get("gene_symbol"))
            evidence = str(target.get("evidence") or "").strip()
            uniprot_id = str(target.get("uniprot_id") or "").strip()
            if symbol is None or evidence not in {"known", "predicted"} or not uniprot_id:
                raise InputError(
                    "drug target entries require gene_symbol, uniprot_id, and valid evidence",
                    details={"ID": compound_id, "target_index": target_index},
                )
            if symbol in seen_symbols:
                raise InputError(
                    "drug_targets.json must be distinct by gene_symbol",
                    details={"ID": compound_id, "gene_symbol": symbol},
                )
            seen_symbols.add(symbol)
            item = dict(target)
            item["gene_symbol"] = symbol
            item["uniprot_id"] = uniprot_id
            item["evidence"] = evidence
            if evidence == "known":
                if predicted_started:
                    raise InputError(
                        "known drug targets must precede predicted targets",
                        details={"ID": compound_id},
                    )
            else:
                predicted_started = True
                rank = target.get("prediction_rank")
                if isinstance(rank, bool) or not isinstance(rank, (int, float)):
                    raise InputError(
                        "predicted targets require numeric prediction_rank",
                        details={"ID": compound_id, "gene_symbol": symbol},
                    )
                integer_rank = int(rank)
                if float(rank) != integer_rank or not 1 <= integer_rank <= 10:
                    raise InputError(
                        "prediction_rank must be an integer from 1 to 10",
                        details={"ID": compound_id, "gene_symbol": symbol},
                    )
                if integer_rank < predicted_rank:
                    raise InputError(
                        "predicted targets must be ordered by prediction_rank",
                        details={"ID": compound_id},
                    )
                predicted_rank = integer_rank
                item["prediction_rank"] = integer_rank
            if "score" in item and item["score"] is not None:
                score = item["score"]
                if isinstance(score, bool) or not isinstance(score, (int, float)):
                    raise InputError(
                        "target score must be numeric when present",
                        details={"ID": compound_id, "gene_symbol": symbol},
                    )
                if not math.isfinite(float(score)):
                    raise InputError("target score must be finite")
                item["score"] = float(score)
            targets.append(item)
        normalized[compound_id] = {
            "smiles": str(info.get("smiles") or ""),
            "targets": targets,
        }
    if not normalized:
        raise InputError("drug_targets.json contains no compounds")
    return normalized


def _implementation_hash() -> str:
    digest = hashlib.sha256()
    digest.update(Path(__file__).read_bytes())
    from . import algorithms

    digest.update(Path(algorithms.__file__).read_bytes())
    digest.update(CACHE_FORMAT_VERSION.encode("utf-8"))
    return digest.hexdigest()


def _load_valid_cache(
    path: Path,
    *,
    cache_key: str,
    node_ids: np.ndarray,
    disease_node_ids: np.ndarray,
    n_random: int,
) -> np.ndarray | None:
    if not path.exists() or not path.is_file() or path.is_symlink():
        return None
    try:
        with np.load(path, allow_pickle=False) as cached:
            required = {
                "node_ids",
                "disease_node_ids",
                "real_distances",
                "random_distances",
                "cache_key",
                "cache_format_version",
            }
            if not required.issubset(cached.files):
                return None
            if str(cached["cache_key"].item()) != cache_key:
                return None
            if str(cached["cache_format_version"].item()) != CACHE_FORMAT_VERSION:
                return None
            if not np.array_equal(cached["node_ids"], node_ids):
                return None
            if not np.array_equal(cached["disease_node_ids"], disease_node_ids):
                return None
            real = np.asarray(cached["real_distances"], dtype=np.int32)
            random_distances = np.asarray(cached["random_distances"], dtype=np.int32)
            if real.shape != (len(node_ids),):
                return None
            if random_distances.shape != (n_random, len(node_ids)):
                return None
            if (real < 0).any() or (random_distances < 0).any():
                return None
            return np.concatenate((real[np.newaxis, :], random_distances), axis=0)
    except (OSError, ValueError, KeyError, EOFError):
        return None


def _cache_arrays(
    matrix: np.ndarray,
    *,
    node_ids: np.ndarray,
    disease_node_ids: np.ndarray,
    cache_key: str,
    code_hash: str,
) -> dict[str, Any]:
    return {
        "node_ids": node_ids,
        "disease_node_ids": disease_node_ids,
        "real_distances": matrix[0],
        "random_distances": matrix[1:],
        "cache_key": np.asarray(cache_key),
        "cache_format_version": np.asarray(CACHE_FORMAT_VERSION),
        "algorithm_version": np.asarray(ALGORITHM_VERSION),
        "code_hash": np.asarray(code_hash),
    }


def _ratio(numerator: int, denominator: int) -> float:
    return 0.0 if denominator == 0 else numerator / denominator


def _operation(
    execution: NodeExecution,
    *,
    context: RunContext,
    ppi_path: str | Path,
    disease_genes_path: str | Path,
    drug_targets_manifest_path: str | Path,
    target_mapping_path: str | Path,
    settings: ProximityConfig,
    shared_cache_dir: str | Path | None,
) -> None:
    started = time.perf_counter()
    execution.update_metrics(
        {
            "n_random": settings.randomizations,
            "minimum_degree_bin_size": settings.minimum_degree_bin_size,
            "seed": settings.seed,
            "job_batch_size": settings.job_batch_size,
            "lower_is_better": settings.lower_is_better,
            "cache_hit": False,
            "cache_source": "miss",
            "elapsed_seconds": 0.0,
        }
    )
    try:
        if settings.background_component != "largest_connected_component":
            raise InputError("proximity background_component must be largest_connected_component")
        if settings.randomization != "degree_matched":
            raise InputError("proximity randomization must be degree_matched")
        if not settings.lower_is_better:
            raise InputError("proximity lower_is_better must remain true")

        ppi_file = resolve_resource_file(context, ppi_path, label="PPI interactome")
        try:
            target_mapping_file = resolve_run_input_file(
                context, target_mapping_path, label="target symbol-to-Entrez mapping"
            )
            mapping_is_resource = False
        except InputError:
            # Retain the standalone CLI's historical fixed-resource mode while
            # allowing the production DAG to consume NetInfer's run artifact.
            target_mapping_file = resolve_resource_file(
                context, target_mapping_path, label="target symbol-to-Entrez mapping"
            )
            mapping_is_resource = True
        disease_file = resolve_run_input_file(
            context, disease_genes_path, label="normalized disease genes"
        )
        drug_targets_file, drug_targets_artifact = resolve_successful_drug_targets(
            context, drug_targets_manifest_path
        )
        execution.input_artifact_ids = tuple(
            dict.fromkeys((*execution.input_artifact_ids, drug_targets_artifact.artifact_id))
        )

        ppi_hash = sha256_file(ppi_file)
        mapping_hash = sha256_file(target_mapping_file)
        disease_hash = sha256_file(disease_file)
        execution.resource_hashes.update({"ppi_interactome": ppi_hash})
        if mapping_is_resource:
            execution.resource_hashes["target_mapping"] = mapping_hash

        raw_graph, ppi_edge_count = _read_ppi(ppi_file)
        largest_component = max(nx.connected_components(raw_graph), key=len)
        graph = raw_graph.subgraph(largest_component).copy()
        node_list = list(graph.nodes())
        node_to_index = {node: index for index, node in enumerate(node_list)}
        disease_genes, disease_input_count, disease_duplicate_count = _read_disease_genes(
            disease_file
        )
        disease_in_ppi = [gene for gene in disease_genes if gene in raw_graph]
        disease_in_lcc = [gene for gene in disease_genes if gene in graph]
        if not disease_in_lcc:
            raise InputError(
                "no normalized disease gene is present in the PPI largest connected component",
                details={
                    "disease_gene_count": len(disease_genes),
                    "ppi_node_count": raw_graph.number_of_nodes(),
                    "lcc_node_count": graph.number_of_nodes(),
                },
            )
        disease_in_lcc.sort(key=node_to_index.__getitem__)

        symbol_to_entrez = _read_target_mapping(target_mapping_file)
        drug_targets = _read_drug_targets(drug_targets_file)
        prepared_rows: list[dict[str, Any]] = []
        skipped_rows: list[dict[str, str]] = []
        unmapped_compounds: dict[str, set[str]] = defaultdict(set)
        target_symbols: set[str] = set()
        mapped_symbols: set[str] = set()
        lcc_symbols: set[str] = set()
        target_entry_count = mapped_entry_count = lcc_entry_count = 0

        for compound_id, info in drug_targets.items():
            targets = info["targets"]
            if not targets:
                skipped_rows.append({"ID": compound_id, "reason": "no_targets"})
                continue
            mapped_entrez: list[str] = []
            lcc_entrez: list[str] = []
            seen_mapped: set[str] = set()
            seen_lcc: set[str] = set()
            symbols: list[str] = []
            for target in targets:
                symbol = target["gene_symbol"]
                symbols.append(symbol)
                target_symbols.add(symbol)
                target_entry_count += 1
                entrez_id = symbol_to_entrez.get(symbol)
                if entrez_id is None:
                    unmapped_compounds[symbol].add(compound_id)
                    continue
                mapped_symbols.add(symbol)
                mapped_entry_count += 1
                if entrez_id not in seen_mapped:
                    mapped_entrez.append(entrez_id)
                    seen_mapped.add(entrez_id)
                if entrez_id in graph:
                    lcc_symbols.add(symbol)
                    lcc_entry_count += 1
                    if entrez_id not in seen_lcc:
                        lcc_entrez.append(entrez_id)
                        seen_lcc.add(entrez_id)
            if not mapped_entrez:
                skipped_rows.append({"ID": compound_id, "reason": "no_mapped_targets"})
                continue
            if not lcc_entrez:
                skipped_rows.append({"ID": compound_id, "reason": "no_targets_in_lcc"})
                continue
            lcc_entrez.sort(key=node_to_index.__getitem__)
            indices = [node_to_index[gene] for gene in lcc_entrez]
            prepared_rows.append(
                {
                    "ID": compound_id,
                    "target_records": json.dumps(
                        targets, ensure_ascii=False, separators=(",", ":")
                    ),
                    "target_gene_symbols": json.dumps(
                        symbols, ensure_ascii=False, separators=(",", ":")
                    ),
                    "target_entrez_ids": json.dumps(mapped_entrez, separators=(",", ":")),
                    "target_lcc_entrez_ids": json.dumps(lcc_entrez, separators=(",", ":")),
                    "target_lcc_indices": json.dumps(indices, separators=(",", ":")),
                    "n_targets_raw": len(targets),
                    "n_targets_mapped": len(mapped_entrez),
                    "n_targets_lcc": len(lcc_entrez),
                }
            )
        if not prepared_rows:
            raise InputError(
                "no compound has a usable target in the PPI largest connected component",
                details={
                    "compound_count": len(drug_targets),
                    "skipped_compound_count": len(skipped_rows),
                },
            )

        unmapped_rows = [
            {
                "gene_symbol": symbol,
                "reason": "missing_entrez_mapping",
                "compound_ids": json.dumps(
                    sorted(compound_ids), ensure_ascii=False, separators=(",", ":")
                ),
            }
            for symbol, compound_ids in sorted(unmapped_compounds.items())
        ]
        unique_target_sets = {tuple(json.loads(row["target_lcc_indices"])) for row in prepared_rows}

        code_hash = _implementation_hash()
        cache_key_payload = {
            "ppi_sha256": ppi_hash,
            "disease_genes_sha256": disease_hash,
            "disease_node_ids": disease_in_lcc,
            "parameters": {
                "randomizations": settings.randomizations,
                "minimum_degree_bin_size": settings.minimum_degree_bin_size,
                "seed": settings.seed,
                "background_component": settings.background_component,
                "randomization": settings.randomization,
            },
            "algorithm_version": ALGORITHM_VERSION,
            "cache_format_version": CACHE_FORMAT_VERSION,
            "code_hash": code_hash,
        }
        cache_key = hash_json(cache_key_payload)
        run_cache_dir, shared_cache = resolve_cache_directories(context, shared_cache_dir)
        cache_filename = f"disease_distance_cache_{cache_key}.npz"
        run_cache_path = run_cache_dir / cache_filename
        node_array = np.asarray(node_list, dtype=str)
        disease_array = np.asarray(disease_in_lcc, dtype=str)

        matrix = _load_valid_cache(
            run_cache_path,
            cache_key=cache_key,
            node_ids=node_array,
            disease_node_ids=disease_array,
            n_random=settings.randomizations,
        )
        cache_source = "run" if matrix is not None else "miss"
        shared_path = None if shared_cache is None else shared_cache / cache_filename
        if matrix is None and shared_path is not None:
            matrix = _load_valid_cache(
                shared_path,
                cache_key=cache_key,
                node_ids=node_array,
                disease_node_ids=disease_array,
                n_random=settings.randomizations,
            )
            if matrix is not None:
                cache_source = "shared"

        # Degree bins/equivalents are cheap relative to the distance matrix and are
        # reconstructed deterministically even on a cache hit for manifest metrics.
        from .algorithms import build_node_to_equivalent_indices, get_degree_binning

        bins = get_degree_binning(graph, settings.minimum_degree_bin_size)
        try:
            build_node_to_equivalent_indices(bins, node_to_index)
        except ValueError as exc:
            raise InputError(
                "PPI is too small for degree-matched randomization",
                details={"minimum_degree_bin_size": settings.minimum_degree_bin_size},
            ) from exc

        distance_execution_mode = "cache"
        if matrix is None:
            try:
                matrix, _, bin_count, distance_execution_mode = build_disease_distance_matrix(
                    graph=graph,
                    node_list=node_list,
                    disease_nodes=disease_in_lcc,
                    n_random=settings.randomizations,
                    minimum_bin_size=settings.minimum_degree_bin_size,
                    seed=settings.seed,
                )
            except ValueError as exc:
                raise InputError(
                    "degree-matched disease randomization could not be constructed",
                    details={"reason": str(exc)},
                ) from exc
        else:
            bin_count = len(bins)

        cache_arrays = _cache_arrays(
            matrix,
            node_ids=node_array,
            disease_node_ids=disease_array,
            cache_key=cache_key,
            code_hash=code_hash,
        )
        # A shared hit is always materialized under run/cache so the artifact remains
        # run-local and the scoring node never depends on mutable external state.
        if cache_source != "run":
            atomic_write_npz(run_cache_path, cache_arrays, allowed_root=run_cache_dir)
        if cache_source == "miss" and shared_path is not None:
            atomic_write_npz(shared_path, cache_arrays, allowed_root=shared_cache)

        network_path = resolve_proximity_output(context, PREPARED_NETWORK_PATH)
        targets_path = resolve_proximity_output(context, PREPARED_TARGETS_PATH)
        skipped_path = resolve_proximity_output(context, SKIPPED_COMPOUNDS_PATH)
        unmapped_path = resolve_proximity_output(context, UNMAPPED_TARGETS_PATH)
        manifest_path = resolve_proximity_output(context, NETWORK_MANIFEST_PATH)
        edges = list(graph.edges())
        atomic_write_npz(
            network_path,
            {
                "node_ids": node_array,
                "edge_source_indices": np.asarray(
                    [node_to_index[left] for left, _ in edges], dtype=np.int32
                ),
                "edge_target_indices": np.asarray(
                    [node_to_index[right] for _, right in edges], dtype=np.int32
                ),
            },
            allowed_root=context.output_dir,
        )
        target_fields = (
            "ID",
            "target_records",
            "target_gene_symbols",
            "target_entrez_ids",
            "target_lcc_entrez_ids",
            "target_lcc_indices",
            "n_targets_raw",
            "n_targets_mapped",
            "n_targets_lcc",
        )
        atomic_write_delimited(
            targets_path,
            prepared_rows,
            fieldnames=target_fields,
            delimiter="\t",
            allowed_root=context.output_dir,
        )
        atomic_write_delimited(
            skipped_path,
            skipped_rows,
            fieldnames=("ID", "reason"),
            delimiter="\t",
            allowed_root=context.output_dir,
        )
        atomic_write_delimited(
            unmapped_path,
            unmapped_rows,
            fieldnames=("gene_symbol", "reason", "compound_ids"),
            delimiter="\t",
            allowed_root=context.output_dir,
        )

        manifest = {
            "schema_version": PREPARATION_SCHEMA_VERSION,
            "algorithm_version": ALGORITHM_VERSION,
            "cache_format_version": CACHE_FORMAT_VERSION,
            "code_hash": code_hash,
            "cache_key": cache_key,
            "cache_key_inputs": cache_key_payload,
            "ppi_resource_sha256": ppi_hash,
            "disease_genes_sha256": disease_hash,
            "target_mapping_sha256": mapping_hash,
            "drug_targets_artifact_id": drug_targets_artifact.artifact_id,
            "lcc_node_count": graph.number_of_nodes(),
            "lcc_edge_count": graph.number_of_edges(),
            "disease_node_count": len(disease_in_lcc),
            "parameters": {
                "randomizations": settings.randomizations,
                "minimum_degree_bin_size": settings.minimum_degree_bin_size,
                "seed": settings.seed,
                "job_batch_size": settings.job_batch_size,
                "lower_is_better": settings.lower_is_better,
            },
            "files": {
                "prepared_network": context.relative_path(network_path),
                "prepared_targets": context.relative_path(targets_path),
                "disease_distance_cache": context.relative_path(run_cache_path),
                "skipped_compounds": context.relative_path(skipped_path),
                "unmapped_targets": context.relative_path(unmapped_path),
            },
            "file_sha256": {
                "prepared_network": sha256_file(network_path),
                "prepared_targets": sha256_file(targets_path),
                "disease_distance_cache": sha256_file(run_cache_path),
                "skipped_compounds": sha256_file(skipped_path),
                "unmapped_targets": sha256_file(unmapped_path),
            },
        }
        atomic_write_json(manifest_path, manifest, allowed_root=context.output_dir)

        metrics = {
            "ppi_input_edge_count": ppi_edge_count,
            "ppi_input_node_count": raw_graph.number_of_nodes(),
            "lcc_node_count": graph.number_of_nodes(),
            "lcc_edge_count": graph.number_of_edges(),
            "disease_input_row_count": disease_input_count,
            "disease_distinct_count": len(disease_genes),
            "disease_duplicate_count": disease_duplicate_count,
            "disease_in_ppi_count": len(disease_in_ppi),
            "disease_in_lcc_count": len(disease_in_lcc),
            "disease_lcc_coverage": _ratio(len(disease_in_lcc), len(disease_genes)),
            "compound_count": len(drug_targets),
            "scorable_compound_count": len(prepared_rows),
            "skipped_compound_count": len(skipped_rows),
            "compound_coverage": _ratio(len(prepared_rows), len(drug_targets)),
            "target_entry_count": target_entry_count,
            "mapped_target_entry_count": mapped_entry_count,
            "lcc_target_entry_count": lcc_entry_count,
            "unique_target_symbol_count": len(target_symbols),
            "unique_mapped_target_symbol_count": len(mapped_symbols),
            "unique_lcc_target_symbol_count": len(lcc_symbols),
            "unmapped_target_symbol_count": len(unmapped_rows),
            "target_mapping_coverage": _ratio(len(mapped_symbols), len(target_symbols)),
            "target_lcc_coverage": _ratio(len(lcc_symbols), len(target_symbols)),
            "unique_target_set_count": len(unique_target_sets),
            "degree_bin_count": bin_count,
            "cache_hit": cache_source != "miss",
            "cache_source": cache_source,
            "cache_key": cache_key,
            "distance_execution_mode": distance_execution_mode,
        }
        execution.update_metrics(metrics)
        if skipped_rows:
            execution.warn(f"Skipped {len(skipped_rows)} compound(s) without usable LCC targets")
        if unmapped_rows:
            execution.warn(f"Could not map {len(unmapped_rows)} unique target symbol(s) to Entrez")
        execution.add_output("proximity_network_manifest", manifest_path)
        execution.add_output("proximity_prepared_network", network_path)
        execution.add_output("proximity_prepared_targets", targets_path)
        execution.add_output("proximity_disease_distance_cache", run_cache_path)
        execution.add_output("proximity_skipped_compounds", skipped_path)
        execution.add_output("proximity_unmapped_targets", unmapped_path)
    finally:
        execution.metric("elapsed_seconds", time.perf_counter() - started)


def proximity_prepare_network(
    *,
    context: RunContext,
    ppi_path: str | Path,
    disease_genes_path: str | Path,
    drug_targets_manifest_path: str | Path,
    target_mapping_path: str | Path,
    settings: ProximityConfig,
    config_hash: str,
    code_version: str,
    shared_cache_dir: str | Path | None = None,
    task_id: str = "main",
    attempt: int = 1,
    input_artifact_ids: Sequence[str] = (),
) -> NodeResult:
    return execute_node(
        lambda execution: _operation(
            execution,
            context=context,
            ppi_path=ppi_path,
            disease_genes_path=disease_genes_path,
            drug_targets_manifest_path=drug_targets_manifest_path,
            target_mapping_path=target_mapping_path,
            settings=settings,
            shared_cache_dir=shared_cache_dir,
        ),
        context=context,
        node_id=NODE_ID,
        task_id=task_id,
        attempt=attempt,
        config_hash=config_hash,
        code_version=code_version,
        input_artifact_ids=input_artifact_ids,
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Prepare PPI LCC, mapped NetInfer targets, and disease-distance cache"
    )
    add_common_runner_arguments(parser)
    parser.add_argument("--ppi", required=True, type=Path)
    parser.add_argument("--disease-genes", required=True, type=Path)
    parser.add_argument("--drug-targets-manifest", required=True, type=Path)
    parser.add_argument("--target-mapping", required=True, type=Path)
    parser.add_argument("--shared-cache-dir", type=Path)
    add_execution_identity_arguments(parser)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    namespace = build_parser().parse_args(argv)
    arguments = CommonRunnerArguments.from_namespace(namespace)
    project_root = Path(__file__).resolve().parents[4]
    environment = load_common_runner_environment(arguments, project_root=project_root)
    task_id, attempt, input_artifact_ids = execution_identity(namespace)
    result = proximity_prepare_network(
        context=environment.context,
        ppi_path=namespace.ppi,
        disease_genes_path=namespace.disease_genes,
        drug_targets_manifest_path=namespace.drug_targets_manifest,
        target_mapping_path=namespace.target_mapping,
        settings=environment.config.proximity,
        config_hash=environment.config_hash,
        code_version=__version__,
        shared_cache_dir=namespace.shared_cache_dir,
        task_id=task_id,
        attempt=attempt,
        input_artifact_ids=input_artifact_ids,
    )
    sys.stdout.write(result.to_json() + "\n")
    return 0 if result.status is NodeStatus.SUCCEEDED else 1


if __name__ == "__main__":
    raise SystemExit(main())


__all__ = [
    "NETWORK_MANIFEST_PATH",
    "PREPARED_NETWORK_PATH",
    "PREPARED_TARGETS_PATH",
    "SKIPPED_COMPOUNDS_PATH",
    "UNMAPPED_TARGETS_PATH",
    "build_parser",
    "main",
    "proximity_prepare_network",
]
